import tkinter as tk
from tkinter import *
from PIL import Image, ImageTk
import json 
import openpyxl 
import time
from datetime import datetime, date, timedelta
from downloadExcel import updateExcelFile
import os

start_time = time.time()


root = tk.Tk()

root.geometry("1920x1080")
root.configure(background='black')
root.attributes('-fullscreen', True)

label = tk.Label(root, text="Bulldog Motorsports ", font=('Arial', 70), bg="black", fg="white")
label.place(x=600, y=50)

MSU_Maroon = "#5D1725"

# Images stored ↓↓↓↓
file_path = "C:/Users/FSAE/Downloads/StatusBoardTest/StatusBoardBulldogMotorsports-main/"
imageLogo = Image.open(file_path + "StatusBoardLogo.png")
logo = ImageTk.PhotoImage(imageLogo)

imageGear = Image.open(file_path + "gearIcon.png")
gear = ImageTk.PhotoImage(imageGear, width=20, height=20)

imageExit = Image.open(file_path + "redX.png")
redX = ImageTk.PhotoImage(imageExit)

imageSave = Image.open(file_path + "saveButton.png")
saveButton = ImageTk.PhotoImage(imageSave)

imageExitMaroon = Image.open(file_path + "maroonX.png")
maroonX = ImageTk.PhotoImage(imageExitMaroon)

imageSaveExit = Image.open(file_path + "save and exit.png")
saveExit = ImageTk.PhotoImage(imageSaveExit)



# Placing Logo


Logo_label = tk.Label(image=logo, borderwidth=0)
Logo_label.image = logo

Logo_label.place(x=400, y=50)


def base_canvas_creator(section_canvas, section_str):
    section_canvas.create_rectangle(0, 0, 300, 600, fill=MSU_Maroon)
    section_canvas.create_text(100, 25, text=section_str, font=("Arial", 25), fill="white")


# Controls Column
controls_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(controls_canvas, "Controls")
controls_canvas.place(x=100, y=200)

# PowerTrain Column
powerTrain_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(powerTrain_canvas, "PowerTrain")
powerTrain_canvas.place(x=350, y=200)

# Drivetrain Column
driveTrain_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(driveTrain_canvas, "DriveTrain")
driveTrain_canvas.place(x=600, y=200)

# Suspension Column
suspension_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(suspension_canvas, "Suspension")
suspension_canvas.place(x=850, y=200)

# Chassis Column
chassis_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(chassis_canvas, "Chassis")
chassis_canvas.place(x=1100, y=200)

# Electrical Column
electrical_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(electrical_canvas, "Electrical")
electrical_canvas.place(x=1350, y=200)

# Aero Column
aero_canvas = tk.Canvas(root, width=200, height=550, borderwidth=2, relief="raised")
base_canvas_creator(aero_canvas, "Aero")
aero_canvas.place(x=1600, y=200)

# Bottom Schedule
footerBar = tk.Canvas(root, width=1500, height=150, borderwidth=2, relief="groove")
footerBar.create_rectangle(0, 0, 1502, 152, fill=MSU_Maroon)
footerBar.place(x=250, y=850)

# Important Dates Text
label = tk.Label(root, text="Important Dates", font=('Arial', 55), bg="black", fg="white")
label.place(x=715, y=760)


# finds the last row of the Excel file
def findLastProjectRow():
    excelOpen = openpyxl.load_workbook(file_path + 'Master Schedule 2024.xlsx')
    activeExcel = excelOpen.active
    for i in range(72, 500):
        cell_section = activeExcel.cell(row=i, column=1)
        if cell_section.value is None:
            break
    return i


# Populates the excelData file
def pullDataFromExcel():
    excelOpen = openpyxl.load_workbook(file_path + 'Master Schedule 2024.xlsx')
    activeExcel = excelOpen.active
    max_row = findLastProjectRow()

    projects = {"Controls": [], "Drivetrain":  [], "Powertrain": [], "Suspension": [], "Chassis": [], "Electrical": [], "Aerodynamics": []}
    for i in range(72, findLastProjectRow()):
        cell_section = activeExcel.cell(row=i, column=1)
        cell_percentage = activeExcel.cell(row=i, column=6)
        cell_date = activeExcel.cell(row=i, column=8)  # end date
        cell_system = activeExcel.cell(row=i, column=2)
        cell_details = activeExcel.cell(row=i, column=4)
        cell_PartStage = activeExcel.cell(row=i, column=3)

        section_value = cell_section.value
        cell_section = section_value.strip()

        # it would make more sense to check the start date to see if the project should be
        # started which is column G in Excel
        # which could use the unused cell_date for a more accurate list of what needs to be started
        if cell_percentage.value != 1.0 and cell_section != "Business":
            if cell_system.value == cell_details.value:
                projects[cell_section].append({"projectname": cell_system.value, "percent": cell_percentage.value, "projectdetails": " ", "date": str(cell_date.value), "Stage": cell_PartStage.value[0]})
            else:
                projects[cell_section].append({"projectname": cell_system.value, "percent": cell_percentage.value, "projectdetails": cell_details.value, "date": str(cell_date.value), "Stage": cell_PartStage.value[0]})
    excelOpen.close()
    return projects


def getProjectData():
    with open(file_path + 'excelData.json', 'r') as json_file:
        projectData = json.load(json_file)
    return projectData


def is_overdue(projectDate):
    try:
        date_format = '%Y-%m-%d'
        projectDate = str(projectDate)
        SlicedProjectDate = projectDate[:10]
        SlicedProjectDate = datetime.strptime(SlicedProjectDate, date_format)
        todays_date = str(date.today())
        SlicedTodaysDate = todays_date[:10]
        SlicedTodaysDate = datetime.strptime(SlicedTodaysDate, date_format)
        if SlicedTodaysDate > SlicedProjectDate + timedelta(days=30):  # Change days for overdue limit
            return True
        else:
            return False
    except:
        print("ERROR: There is something wrong in the excel file")
    

def cutStringToLength(text):
    sliced = ""
    if len(text) > 17:  # if its oversized
        for i in reversed(text):  # start at end and go through 1 by 1
            if i == " ":  # if i's character == " " then
                break
            sliced = i + sliced
        return sliced
    else: 
        return text


# We can only display 33 characters in between projectname and project details
def displayData(section_canvas, section_str):
    projectData = getProjectData()

    g = 0

    for i in range(7):
        text = f'{projectData[section_str][i]["projectname"]} {projectData[section_str][i]["projectdetails"]}'
        limitText = f'{cutStringToLength(text)} ({projectData[section_str][i]["Stage"]})'
        section_canvas.create_text(40, 75 + g, text=limitText, font=("Arial", 12), fill="white", anchor='w')
        percentage = projectData[section_str][i]["percent"]
        section_canvas.create_rectangle(40, 88 + g, 180, 108 + g, fill="white")
        if percentage == 0.00:
            section_canvas.create_rectangle(40, 88 + g, 40, 108 + g, fill="orange")
        elif is_overdue(projectData[section_str][i]["date"]) and percentage != 0.00:
            section_canvas.create_rectangle(40, 88 + g, 40 + (140 * percentage), 108 + g, fill="red")
        else:
            section_canvas.create_rectangle(40, 88 + g, 40 + (140 * percentage), 108 + g, fill="orange")
        section_canvas.create_text(130, 98 + g, text=f'{int(projectData[section_str][i]["percent"] * 100)}%', font=("Arial bold", 12), fill="black", anchor='e')
        g += 70


# Button that calls edit function for important Dates (CONTAINS MANY FUNCTIONS TO
# READ AND WRITE TO JSON AND TO DISPLAY DATES)


def editButton():
    edit_window = tk.Toplevel(root)
    edit_window.title("Edit Important Dates")
    edit_window.geometry("675x250+70+70")
    edit_window.overrideredirect(True)
    edit_window.configure(background=MSU_Maroon, borderwidth=7, relief="solid", highlightthickness=3)

    # Text Entry Boxes Defined and placed below

    task1 = tk.Label(edit_window, text="Task one:")
    task1Task = tk.Label(edit_window, text="Task:")
    task1Date = tk.Label(edit_window, text="Date:")
    task1Entry = tk.Entry(edit_window)
    task1EntryDate = tk.Entry(edit_window)
    
    task2 = tk.Label(edit_window, text="Task two:")
    task2Task = tk.Label(edit_window, text="Task:")
    task2Date = tk.Label(edit_window, text="Date:")
    task2Entry = tk.Entry(edit_window)
    task2EntryDate = tk.Entry(edit_window)
    
    task3 = tk.Label(edit_window, text="Task three:")
    task3Task = tk.Label(edit_window, text="Task:")
    task3Date = tk.Label(edit_window, text="Date:")
    task3Entry = tk.Entry(edit_window)
    task3EntryDate = tk.Entry(edit_window)
    
    task4 = tk.Label(edit_window, text="Task four:")
    task4Task = tk.Label(edit_window, text="Task:")
    task4Date = tk.Label(edit_window, text="Date:")
    task4Entry = tk.Entry(edit_window)
    task4EntryDate = tk.Entry(edit_window)
    
    task5 = tk.Label(edit_window, text="Task five:")
    task5Task = tk.Label(edit_window, text="Task:")
    task5Date = tk.Label(edit_window, text="Date:")
    task5Entry = tk.Entry(edit_window)
    task5EntryDate = tk.Entry(edit_window)

    task6 = tk.Label(edit_window, text="Task six:")
    task6Task = tk.Label(edit_window, text="Task:")
    task6Date = tk.Label(edit_window, text="Date:")
    task6Entry = tk.Entry(edit_window)
    task6EntryDate = tk.Entry(edit_window)

    # Sets the data from the entries to a dictionary

    def setImportantDates():
        data = {
            "task1": {
                "task": task1Entry.get(),
                "date": task1EntryDate.get()
            },
            "task2": {
                "task": task2Entry.get(),
                "date": task2EntryDate.get()
            },
            "task3": {
                "task": task3Entry.get(),
                "date": task3EntryDate.get()
            },
            "task4": {
                "task": task4Entry.get(),
                "date": task4EntryDate.get()
            },
            "task5": {
                "task": task5Entry.get(),
                "date": task5EntryDate.get()
            },
            "task6": {
                "task": task6Entry.get(),
                "date": task6EntryDate.get()
            }
        }
        # writes it to important_dates.json
        with open(file_path + "important_dates.json", "w") as json_file:
            json.dump(data, json_file)

    # reads the data that was put into important_dates and returns the data in the form of a dictionary

    # places the data from the dictionary into displayable text

    # will update the data on closure of the edit window
    def exit_editButton():
        setImportantDates()
        displayImportantDates()
        edit_window.destroy()

    # save_button = tk.Button(edit_window, text="Save", image=saveButton, borderwidth=0, highlightthickness=0, command=setImportantDates)
    exitSave_button = tk.Button(edit_window, text="Exit", command=exit_editButton, image=saveExit, borderwidth=0, highlightthickness=0)
    exit_button = tk.Button(edit_window, text="Exit", command=edit_window.destroy, image=maroonX, borderwidth=0, highlightthickness=0)

    # Placing text boxes for important dates entry
    task1.place(x=70, y=10)
    task1Entry.place(x=50, y=35)
    task1Task.place(x=18, y=35)
    task1Date.place(x=18, y=60)
    task1EntryDate.place(x=50, y=60)

    task2.place(x=295, y=10)
    task2Entry.place(x=275, y=35)
    task2Task.place(x=243, y=35)
    task2Date.place(x=243, y=60)
    task2EntryDate.place(x=275, y=60)

    task3.place(x=525, y=10)
    task3Entry.place(x=500, y=35)
    task3Task.place(x=468, y=35)
    task3Date.place(x=468, y=60)
    task3EntryDate.place(x=500, y=60)

    task4.place(x=70, y=95)
    task4Entry.place(x=50, y=120)
    task4Task.place(x=18, y=120)
    task4Date.place(x=18, y=145)
    task4EntryDate.place(x=50, y=145)

    task5.place(x=295, y=95)
    task5Entry.place(x=275, y=120)
    task5Task.place(x=243, y=120)
    task5Date.place(x=243, y=145)
    task5EntryDate.place(x=275, y=145)

    task6.place(x=530, y=95)
    task6Entry.place(x=500, y=120)
    task6Task.place(x=468, y=120)
    task6Date.place(x=468, y=145)
    task6EntryDate.place(x=500, y=145)

    exitSave_button.place(x=220, y=175)
    exit_button.place(x=620, y=0)


# Button to access edit function
ImpDatesButton = tk.Button(text="Edit Important Dates", command=editButton, image=gear, borderwidth=0, highlightthickness=0)

ImpDatesButton.place(x=85, y=20)

# Display without pressing edit   HAS ISSUES BECAUSE IT OVERLAPS EDITED DATA


def getImportantDates():
    with open(file_path + "important_dates.json", "r") as json_file:
        importantDataList = json.load(json_file)
    return importantDataList


def displayImportantDates():
    data = getImportantDates()
    footerBar.create_rectangle(0, 0, 1502, 152, fill=MSU_Maroon)  # Drawing over the old maroon rectangle
    b = 0
    for i in range(0, 6):
        key = f'task{i + 1}'
        if f'{data[key]["task"]}' != "":
            x, y = [(50, 40), (550, 40), (1050, 40), (50, 110), (550, 110), (1050, 110)][b]
            footerBar.create_text(x, y, text=f'• {data[key]["task"]}  {data[key]["date"]}', font=("Helvetica", 20), fill="white", anchor="w")
            b += 1


# Update's Excel File

def dumpProjectData2JSON():
    projectData = pullDataFromExcel()
    with open(file_path + 'excelData.json', 'w') as json_file:
        json.dump(projectData, json_file)


updateExcelFile()
dumpProjectData2JSON()


displayImportantDates()


displayData(controls_canvas, "Controls")
displayData(powerTrain_canvas, "Powertrain")
displayData(driveTrain_canvas, "Drivetrain")
displayData(suspension_canvas, "Suspension")
displayData(chassis_canvas, "Chassis")
displayData(electrical_canvas, "Electrical")
displayData(aero_canvas, "Aerodynamics")

# Exit button for entire program

ExitButton = tk.Button(text="Exit", command=root.destroy, image=redX, borderwidth=0, highlightthickness=0)

ExitButton.place(x=1830, y=30)

print("--- %s seconds ---" % (time.time() - start_time))


root.mainloop()
